import re
import sqlite3

from bs4 import BeautifulSoup
from os import path, makedirs

import time

import requests
import json
from datetime import date

import openpyxl
from openpyxl.styles import PatternFill, Font

COLOR = "00FF00"
headers = {
    "Accept": "*/*",
    "Content-Type": "text/plain",
    "Origin": "https://www.zacks.com",
    "Accept-Language": "ru",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.6 Safari/605.1.15",
    "Referer": "https://www.zacks.com/",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive"
}


class Parser:
    def bm(self, num):
        if num is None or num == "NA" or num == "N/A":
            return num
        if "M" in num:
            return float(num.replace("M", "")) * 1000000
        elif "B" in num:
            return float(num.replace("B", "")) * 1000000000
        else:
            return float(num)

    def __init__(self, symbol, timeout):
        self.symbol = symbol

        self.url_Calendar = f"https://www.zacks.com/stock/research/{symbol}/earnings-calendar"
        # self.url_Consensus = f"https://www.zacks.com//data_handler/charts/?ticker={symbol}&wrapper=price_and_eps_estimates_consensus&addl_settings=time_period=2019,2020,2021,2022,2023,2024,2025"
        self.url_Consensus = f"https://www.zacks.com//data_handler/charts/?ticker={symbol}&wrapper=price_and_eps_estimates_consensus&addl_settings=time_period=2016,2017,2018,2019,2020,2021,2022,2023,2024,2025"
        self.url_Detailed = f"https://www.zacks.com/stock/quote/{symbol}/detailed-earning-estimates"

        self.data = {
            "symbol": symbol,
            "json_data": {
                "Calendar": {},
                "Consensus": {},
                "Detailed": {}
            }
        }
        self.timeout = timeout

        print(symbol, end=" => ")

        self.get_Detailed()
        time.sleep(self.timeout)

        self.get_Calendar()
        time.sleep(self.timeout)

        self.get_Consensus()
        time.sleep(self.timeout)

    def get_Detailed(self):
        try:
            print("det,", end=" ")

            data = self.data["json_data"]["Detailed"]
            try:
                response = requests.get(self.url_Detailed, headers=headers, timeout=10)
            except requests.exceptions.ConnectTimeout as e:
                print("/restart det/,", end=" ")
                time.sleep(10)
                try:
                    response = requests.get(self.url_Detailed, headers=headers, timeout=10)
                except Exception as e:
                    print()
                    data["Sales_Cur_high"] = None
                    data["Sales_Cur_low"] = None
                    data["Sales_Next_high"] = None
                    data["Sales_Next_low"] = None
                    data["Ear_Cur_high"] = None
                    data["Ear_Cur_low"] = None
                    data["Ear_Next_high"] = None
                    data["Ear_Next_low"] = None

            # with open(f"HTML/{self.symbol}_detailed.html", "w") as f:
            #     f.write(response.text)

            soup = BeautifulSoup(response.text, "html.parser")
            try:
                table_Sales = soup.find("h1", text="Sales Estimates").parent.find("tbody")
                S_high = table_Sales.find("td", text="High Estimate").parent.find_all("td")
                S_low = table_Sales.find("td", text="Low Estimate").parent.find_all("td")

                table_Ear = soup.find("h1", text="Earnings Estimates").parent.find("tbody")
                E_high = table_Ear.find("td", text="High Estimate").parent.find_all("td")
                E_low = table_Ear.find("td", text="Low Estimate").parent.find_all("td")

                data["Sales_Cur_high"] = self.bm(S_high[3].text)
                data["Sales_Cur_low"] = self.bm(S_low[3].text)
                data["Sales_Next_high"] = self.bm(S_high[4].text)
                data["Sales_Next_low"] = self.bm(S_low[4].text)

                data["Ear_Cur_high"] = self.bm(E_high[3].text)
                data["Ear_Cur_low"] = self.bm(E_low[3].text)
                data["Ear_Next_high"] = self.bm(E_high[4].text)
                data["Ear_Next_low"] = self.bm(E_low[4].text)
            except AttributeError:
                data["Sales_Cur_high"] = None
                data["Sales_Cur_low"] = None
                data["Sales_Next_high"] = None
                data["Sales_Next_low"] = None

                data["Ear_Cur_high"] = None
                data["Ear_Cur_low"] = None
                data["Ear_Next_high"] = None
                data["Ear_Next_low"] = None
        except Exception as e:
            data = self.data["json_data"]["Detailed"]
            data["Sales_Cur_high"] = None
            data["Sales_Cur_low"] = None
            data["Sales_Next_high"] = None
            data["Sales_Next_low"] = None
            data["Ear_Cur_high"] = None
            data["Ear_Cur_low"] = None
            data["Ear_Next_high"] = None
            data["Ear_Next_low"] = None

    def get_Calendar(self):
        try:
            print("cal,", end=" ")

            data = self.data["json_data"]["Calendar"]

            response = requests.get(self.url_Calendar, headers=headers, timeout=10)

            # with open(f"HTML/{self.symbol}_calendar.html", "w") as f:
            #     f.write(response.text)

            soup = BeautifulSoup(response.text, "html.parser")

            pattern = re.compile(r'obj_data')
            try:
                script_data = soup.find("script", text=pattern).contents[0]
            except AttributeError:
                data["Ear"] = []
                return
            d = script_data.find('{')
            e = script_data.find('}')

            json_data = json.loads(script_data[d:e + 1])
            # with open("WMT.json", "w") as f:
            #     json.dump(json_data, f)

            ear = json_data["earnings_announcements_earnings_table"]
            for spis in ear:
                spis[4] = spis[4][spis[4].find('">') + 2:spis[4].find('</div>')]
                spis[5] = spis[5][spis[5].find('">') + 2:spis[5].find('</div>')]
            data["Ear"] = ear

            # with open(f"results_json/{symbol}_calendar.json", "w") as f:
            #     json.dump(data, f)
        except Exception as e:
            data = self.data["json_data"]["Calendar"]
            data["Ear"] = []

    def get_Consensus(self):
        try:
            print("con")
            response = requests.get(self.url_Consensus, headers=headers, timeout=10)

            if response.ok:
                dt = response.json()

                for i in range(2016, 2026):
                    self.data["json_data"]["Consensus"][i] = dt.get(f"eps_consensus_{i}")
        except Exception as e:
            self.data["json_data"]["Consensus"] = {
                "2016": {},
                "2017": {},
                "2018": {},
                "2019": {},
                "2020": {},
                "2021": {},
                "2022": {},
                "2023": {},
                "2024": {},
                "2025": {}
            }


class Xlsx:
    def cell_header(self, cell, text, color=False):
        cell.value = text
        cell.font = Font(bold=True)
        # cell.alignment = Alignment(horizontal="center")

        if color:
            cell.fill = self.redFill

    def __init__(self, folder):
        self.folder = folder
        self.years = [str(y) for y in range(2016, 2026)]

        self.cal_file = 0

        self.redFill = PatternFill(start_color=COLOR, fill_type='solid')
        self.xlsx_main()

        self.row_Detailed = 5
        self.cell_Calendar = -1
        self.cell_Consensus = 1

    def xlsx_main(self):
        book = openpyxl.Workbook()

        book.save(f"{self.folder}/Calendar_{str(self.cal_file)}.xlsx")
        book.close()

        book = openpyxl.Workbook()

        book.save(f"{self.folder}/Consensus_{str(self.cal_file)}.xlsx")
        book.close()

        print("Xlsx files are created")
        print()

    def to_xlsx(self, data):
        try:
            self.to_xlsx_Calendar(data)
            self.cell_Calendar += 3
        except Exception as e:
            print(f"Data Calendar is not saved in xlsx file, {e}")
        try:
            self.to_xlsx_Detailed(data)
            self.row_Detailed += 1
        except Exception as e:
            print(f"Data Detailed is not saved in xlsx file, {e}")

        try:
            self.to_xlsx_Consensus(data)
            self.cell_Consensus += (len(self.years) * 2)
        except Exception as e:
            print(f"Data Consensus is not saved in xlsx file, {e}")

    def to_xlsx_Detailed(self, data):
        book = openpyxl.open("RESULTS/result.xlsx")
        sheet = book.active
        shift = 130

        # sheet.cell(row=self.row_Detailed, column=shift + 1).value = data["symbol"]

        sheet.cell(row=self.row_Detailed, column=shift + 1).value = data["json_data"]["Detailed"].get("Sales_Cur_high")
        sheet.cell(row=self.row_Detailed, column=shift + 2).value = data["json_data"]["Detailed"].get("Sales_Cur_low")
        sheet.cell(row=self.row_Detailed, column=shift + 3).value = data["json_data"]["Detailed"].get("Sales_Next_high")
        sheet.cell(row=self.row_Detailed, column=shift + 4).value = data["json_data"]["Detailed"].get("Sales_Next_low")

        sheet.cell(row=self.row_Detailed, column=shift + 5).value = data["json_data"]["Detailed"].get("Ear_Cur_high")
        sheet.cell(row=self.row_Detailed, column=shift + 6).value = data["json_data"]["Detailed"].get("Ear_Cur_low")
        sheet.cell(row=self.row_Detailed, column=shift + 7).value = data["json_data"]["Detailed"].get("Ear_Next_high")
        sheet.cell(row=self.row_Detailed, column=shift + 8).value = data["json_data"]["Detailed"].get("Ear_Next_low")

        book.save("RESULTS/result.xlsx")
        book.close()

    def to_xlsx_Calendar(self, data):
        book = openpyxl.open(f"{self.folder}/Calendar_{str(self.cal_file)}.xlsx")
        sheet = book.active

        self.cell_header(sheet.cell(row=1, column=self.cell_Calendar + 2), data["symbol"], color=True)

        self.cell_header(sheet.cell(row=2, column=self.cell_Calendar + 2), "Earnings")

        self.cell_header(sheet.cell(row=3, column=self.cell_Calendar + 2), "Date")
        self.cell_header(sheet.cell(row=3, column=self.cell_Calendar + 3), "% Surprise")
        self.cell_header(sheet.cell(row=3, column=self.cell_Calendar + 4), "Time")

        book.save(f"{self.folder}/Calendar_{str(self.cal_file)}.xlsx")

        ear = data["json_data"]["Calendar"].get("Ear")
        if ear:
            for i in range(len(ear)):
                try:
                    dd = ear[i][0].split("/")
                    sheet.cell(row=4 + i, column=self.cell_Calendar + 2).value = date(
                        int("20" + dd[2]), int(dd[0]), int(dd[1]))
                except Exception:
                    sheet.cell(row=4 + i, column=self.cell_Calendar + 2).value = ear[i][0]
                sheet.cell(row=4 + i, column=self.cell_Calendar + 3).value = ear[i][5]
                sheet.cell(row=4 + i, column=self.cell_Calendar + 4).value = ear[i][6]

            book.save(f"{self.folder}/Calendar_{str(self.cal_file)}.xlsx")
        book.save(f"{self.folder}/Calendar_{str(self.cal_file)}.xlsx")
        book.close()

    def to_xlsx_Consensus(self, data):
        book = openpyxl.open(f"{self.folder}/Consensus_{str(self.cal_file)}.xlsx")
        sheet = book.active

        self.cell_header(sheet.cell(row=1, column=self.cell_Consensus), data["symbol"], color=True)

        for m, year in enumerate(self.years):
            self.cell_header(sheet.cell(row=2, column=self.cell_Consensus + m * 2), year)

        for i in range(len(self.years)):
            self.cell_header(sheet.cell(row=3, column=self.cell_Consensus + 2 * i), "Date")
            self.cell_header(sheet.cell(row=3, column=self.cell_Consensus + 1 + 2 * i), "Value")
        book.save(f"{self.folder}/Consensus_{str(self.cal_file)}.xlsx")

        # -=-=-=-=-=-==
        json_data = data["json_data"]["Consensus"]

        for i, year in enumerate(self.years):
            dt = json_data[int(year)]
            if dt:
                for row, key in enumerate(dt.keys()):
                    dd = key.split("/")
                    sheet.cell(row=row + 4, column=self.cell_Consensus + 2 * i).value = date(
                        int("20" + dd[2]), int(dd[0]), int(dd[1]))
                    sheet.cell(row=row + 4, column=self.cell_Consensus + 1 + 2 * i).value = float(dt.get(key, "0.0").replace("N/A", "0.0"))
                book.save(f"{self.folder}/Consensus_{str(self.cal_file)}.xlsx")

        book.save(f"{self.folder}/Consensus_{str(self.cal_file)}.xlsx")
        book.close()


def del_all():
    con = sqlite3.connect("all_analyst.db")
    cur = con.cursor()
    cur.execute("DELETE FROM zacks_detailed")
    cur.execute("DELETE FROM zacks_consensus")
    cur.execute("DELETE FROM zacks_calendar")
    con.commit()
    con.close()


def save_data(data):
    ticker = data.get("symbol")
    cal = data.get("json_data").get("Calendar").get("Ear", [])
    cons = data.get("json_data").get("Consensus")
    det = data.get("json_data").get("Detailed")

    con = sqlite3.connect("all_analyst.db")
    cur = con.cursor()

    r = f"""INSERT INTO zacks_detailed VALUES
                ({'?' + ',?' * 8});"""

    cur.execute(r, (ticker, det.get("Sales_Cur_high"), det.get("Sales_Cur_low"), det.get("Sales_Next_high"), det.get("Sales_Next_low"), det.get("Ear_Cur_high"), det.get("Ear_Cur_low"), det.get("Ear_Next_high"), det.get("Ear_Next_low")))
    con.commit()

    r = f"""INSERT INTO zacks_consensus VALUES
                    (?, ?, ?);"""

    for key in cons:
        dt = cons[key]
        if dt:
            for k in dt:
                cur.execute(r, (ticker, k, dt[k]))
            con.commit()

    r = f"""INSERT INTO zacks_calendar VALUES
                (?, ?, ?, ?);"""
    for cort in cal:
        cur.execute(r, (ticker, cort[0], cort[-2], cort[-1]))
        con.commit()
    con.close()


def main(appl_data):
    # if True:
    try:
        print()
        print("-=-=- ZACKS -=-=-")

        with open("config.txt") as f:
            data = f.readlines()
            ZACKS_TIMEOUT = data[5].strip().split("=")[1].replace('"', '')
        # print([ZACKS_TIMEOUT])

        folder = f"RESULTS/ZACKS_RES"
        if not path.exists(folder):
            makedirs(folder)

        xlsx = Xlsx(folder)

        del_all()
        for symbol in appl_data:
            inform = Parser(symbol, int(ZACKS_TIMEOUT))
            # with open(f"zacks/JSON/{symbol}_result.json", "w") as f:
            #     json.dump(inform.data, f)
            save_data(inform.data)
            xlsx.to_xlsx(inform.data)

            # with open(f"zacks/JSON/{symbol}_result.json") as f:
            #     data = json.load(f)
            #     save_data(data)
            #     xlsx.to_xlsx(data)
    except Exception:
        print("BIG ERROR ZACKS")
