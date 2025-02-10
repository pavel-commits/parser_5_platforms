import openpyxl
import requests
import re
import time
import json
# import os.path
from bs4 import BeautifulSoup
from datetime import date
# from yfinance import Ticker
import warnings
# from pandas import Timestamp
import sqlite3
from openpyxl.styles import PatternFill, Font

COLOR = "00FF00"


headers = {
    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.3 Safari/605.1.15"
}

warnings.filterwarnings("ignore")
COO = 10


def cell_header(cell, text, color=False):
    cell.value = text
    cell.font = Font(bold=True)

    if color:
        cell.fill = PatternFill(start_color=COLOR, fill_type='solid')


class Parser:
    def __init__(self, AAPL, timeout, cookie, crumb):
        self.AAPL = AAPL
        self.timeout = timeout
        self.cookie = cookie

        self.url_Corporate = f"https://finance.yahoo.com/quote/{AAPL}/profile?p={AAPL}"
        self.url_Api = f"https://query2.finance.yahoo.com/v10/finance/quoteSummary/{AAPL}?crumb={crumb}&modules=assetProfile%2CesgScores%2CfinancialData%2CdefaultKeyStatistics%2CsummaryDetail%2CearningsHistory%2CearningsTrend"

        # self.url_Api = f"https://query2.finance.yahoo.com/v10/finance/quoteSummary/{AAPL}?modules=assetProfile%2CsummaryProfile%2CsummaryDetail%2CesgScores%2Cprice%2CincomeStatementHistory%2CincomeStatementHistoryQuarterly%2CbalanceSheetHistory%2CbalanceSheetHistoryQuarterly%2CcashflowStatementHistory%2CcashflowStatementHistoryQuarterly%2CdefaultKeyStatistics%2CfinancialData%2CcalendarEvents%2CsecFilings%2CrecommendationTrend%2CupgradeDowngradeHistory%2CinstitutionOwnership%2CfundOwnership%2CmajorDirectHolders%2CmajorHoldersBreakdown%2CinsiderTransactions%2CinsiderHolders%2CnetSharePurchaseActivity%2Cearnings%2CearningsHistory%2CearningsTrend%2CindustryTrend%2CindexTrend%2CsectorTrend"
        # self.url_Api = f"https://query2.finance.yahoo.com/v10/finance/quoteSummary/{AAPL}?modules=assetProfile%2CesgScores%2CfinancialData%2CdefaultKeyStatistics%2CsummaryDetail%2CearningsHistory%2CearningsTrend"

        self.data = {
            "Summary": {
                "1yTargetEst": None
            },
            "Profile": {
                "fullTimeEmployees": None,
                "corporateGovernance": None
            },
            "Statistics": {
                "averageDailyVolume3Month": None,
                "shares_short": None,
                "shortPercentOfFloat": None,
                "sharesShortPriorMonth": None,
                "trailingAnnualDividendRate": None
            },
            "Analysis": {
                "numberOfAnalysts": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "earnings_avg": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "low": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "high": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "revenue_avg": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "salesGrowth": {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None},
                "epsEstimate": None,
                "epsActual": None,
                "currentEstimate": None,
                "7daysAgo": None,
                "30daysAgo": None,
                "currentYear": None,
                "nextYear": None,
                "next5Years": None,
                "past5Years": None
            },
            "Sustainability": {
                "totalEsg": None,
                "environmentScore": None,
                "socialScore": None,
                "governanceScore": None,
                "highestControversy": None
            },
        }
        print(self.AAPL, end=" ")

        mas = self.get_Api()
        if not mas:
            self.get_corporate_governanse()

    def get_Api(self):
        # if True:
        try:
            print("1", end=" ")
            response = requests.get(
                self.url_Api,
                headers=headers,
                cookies={self.cookie.name: self.cookie.value},
                allow_redirects=True
            )

            if response:
                response = response.json()
                # with open("1.json", "w") as fdf:
                #     json.dump(response, fdf)
                if response.get("quoteSummary", {}).get("error") is not None:
                    print(f"No api data")
                    return "NONE"
            else:
                print(f"No api data")
                return "NONE"
            json_data = response.get("quoteSummary").get("result")[0]

            # "Profile" & "Sustainability"
            pr = json_data.get("assetProfile", {})

            self.data["Profile"]["fullTimeEmployees"] = pr.get("fullTimeEmployees", None)
            sus = {
                key: value for key, value in json_data.get("esgScores", {}).items()
                if value is not None
            }

            self.data["Sustainability"]["totalEsg"] = sus.get("totalEsg", {}).get("raw", None)
            self.data["Sustainability"]["environmentScore"] = sus.get("environmentScore", {}).get("raw", None)
            self.data["Sustainability"]["socialScore"] = sus.get("socialScore", {}).get("raw", None)
            self.data["Sustainability"]["governanceScore"] = sus.get("governanceScore", {}).get("raw", None)
            self.data["Sustainability"]["highestControversy"] = sus.get("highestControversy", None)

            # "Summary" & "Statistics"
            self.data["Summary"]["1yTargetEst"] = json_data.get("financialData", {}).get("targetMeanPrice", {}).get(
                "raw", None)

            sd = {
                key: value for key, value in json_data.get("summaryDetail", {}).items()
                if value is not None
            }
            self.data["Statistics"]["averageDailyVolume3Month"] = sd.get("averageVolume", {}).get("raw", None)
            self.data["Statistics"]["trailingAnnualDividendRate"] = sd.get("trailingAnnualDividendRate", {}).get(
                "raw",
                None)

            dks = {
                key: value for key, value in json_data.get("defaultKeyStatistics", {}).items()
                if value is not None
            }
            self.data["Statistics"]["sharesShortPriorMonth"] = dks.get("sharesShortPriorMonth", {}).get("raw", None)
            self.data["Statistics"]["shares_short"] = dks.get("sharesShort", {}).get("raw", None)
            self.data["Statistics"]["shortPercentOfFloat"] = dks.get("shortPercentOfFloat", {}).get("raw", None)

            # Analysys
            tr = json_data.get("earningsTrend", {}).get("trend", [{}, {}, {}, {}, {}, {}])
            trend_0 = tr[0]
            trend_1 = tr[1]
            trend_2 = tr[2]
            trend_3 = tr[3]

            # =--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
            ear_0 = trend_0.get("earningsEstimate", {})
            ear_1 = trend_1.get("earningsEstimate", {})
            ear_2 = trend_2.get("earningsEstimate", {})
            ear_3 = trend_3.get("earningsEstimate", {})

            # {"CurrentQtr": None, "NextQtr": None, "CurrentYear": None, "NextYear": None}
            self.data["Analysis"]["numberOfAnalysts"]["CurrentYear"] = ear_2.get("numberOfAnalysts", {}).get("raw",
                                                                                                             None)
            self.data["Analysis"]["numberOfAnalysts"]["NextYear"] = ear_3.get("numberOfAnalysts", {}).get("raw",
                                                                                                          None)
            self.data["Analysis"]["numberOfAnalysts"]["CurrentQtr"] = ear_0.get("numberOfAnalysts", {}).get("raw",
                                                                                                            None)
            self.data["Analysis"]["numberOfAnalysts"]["NextQtr"] = ear_1.get("numberOfAnalysts", {}).get("raw",
                                                                                                         None)

            self.data["Analysis"]["earnings_avg"]["CurrentYear"] = ear_2.get("avg", {}).get("raw", None)
            self.data["Analysis"]["earnings_avg"]["NextYear"] = ear_3.get("avg", {}).get("raw", None)
            self.data["Analysis"]["earnings_avg"]["CurrentQtr"] = ear_0.get("avg", {}).get("raw", None)
            self.data["Analysis"]["earnings_avg"]["NextQtr"] = ear_1.get("avg", {}).get("raw", None)

            self.data["Analysis"]["low"]["CurrentYear"] = ear_2.get("low", {}).get("raw", None)
            self.data["Analysis"]["low"]["NextYear"] = ear_3.get("low", {}).get("raw", None)
            self.data["Analysis"]["low"]["CurrentQtr"] = ear_0.get("low", {}).get("raw", None)
            self.data["Analysis"]["low"]["NextQtr"] = ear_1.get("low", {}).get("raw", None)

            self.data["Analysis"]["high"]["CurrentYear"] = ear_2.get("high", {}).get("raw", None)
            self.data["Analysis"]["high"]["NextYear"] = ear_3.get("high", {}).get("raw", None)
            self.data["Analysis"]["high"]["CurrentQtr"] = ear_0.get("high", {}).get("raw", None)
            self.data["Analysis"]["high"]["NextQtr"] = ear_1.get("high", {}).get("raw", None)

            # =--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
            rev_0 = trend_0.get("revenueEstimate", {})
            rev_1 = trend_1.get("revenueEstimate", {})
            rev_2 = trend_2.get("revenueEstimate", {})
            rev_3 = trend_3.get("revenueEstimate", {})

            self.data["Analysis"]["revenue_avg"]["CurrentYear"] = rev_2.get("avg", {}).get("raw")
            self.data["Analysis"]["revenue_avg"]["NextYear"] = rev_3.get("avg", {}).get("raw")
            self.data["Analysis"]["revenue_avg"]["CurrentQtr"] = rev_0.get("avg", {}).get("raw")
            self.data["Analysis"]["revenue_avg"]["NextQtr"] = rev_1.get("avg", {}).get("raw")

            self.data["Analysis"]["salesGrowth"]["CurrentYear"] = rev_2.get("growth", {}).get("raw")
            self.data["Analysis"]["salesGrowth"]["NextYear"] = rev_3.get("growth", {}).get("raw")
            self.data["Analysis"]["salesGrowth"]["CurrentQtr"] = rev_0.get("growth", {}).get("raw")
            self.data["Analysis"]["salesGrowth"]["NextQtr"] = rev_1.get("growth", {}).get("raw")

            # =--=-=-=-=-=-=-=-=-=-=-=-=-=-=-
            hist = json_data.get("earningsHistory", {}).get("history", [{}])[0]
            self.data["Analysis"]["epsEstimate"] = hist.get("epsEstimate", {}).get("raw", None)
            self.data["Analysis"]["epsActual"] = hist.get("epsActual", {}).get("raw", None)

            # =--=-=-=-=-=-=-=-=-=-=-=-=-=-=-
            eps_0 = trend_0.get("epsTrend", {})

            self.data["Analysis"]["currentEstimate"] = eps_0.get("current", {}).get("raw", None)
            self.data["Analysis"]["7daysAgo"] = eps_0.get("7daysAgo", {}).get("raw", None)
            self.data["Analysis"]["30daysAgo"] = eps_0.get("30daysAgo", {}).get("raw", None)

            # =--=-=-=-=-=-=-=-=-=-=-=-=-=-=-

            self.data["Analysis"]["currentYear"] = tr[2].get("growth", {}).get("raw", None)
            self.data["Analysis"]["nextYear"] = tr[3].get("growth", {}).get("raw", None)
            self.data["Analysis"]["next5Years"] = tr[4].get("growth", {}).get("raw", None)
            self.data["Analysis"]["past5Years"] = tr[5].get("growth", {}).get("raw", None)
        except Exception as e:
            print(f"error get_api {self.url_Api}")
            print(e)

    def get_corporate_governanse(self):
        try:
            print("2")
            response = requests.get(self.url_Corporate, headers=headers)

            soup = BeautifulSoup(response.text, "html.parser")
            container = soup.find("section", class_="corporate-governance-container").find("div",
                                                                                           class_="Mt(20px)").find(
                "span", text=re.compile(r'while')).text
            container = container[container.find("while a "):].split()
            self.data["Profile"]["corporateGovernance"] = int(container[2])
        except Exception as e:
            print("error corporate_governanse", self.url_Corporate)


def to_xlsx(data, row):
    book = openpyxl.open("RESULTS/result.xlsx")
    sheet = book.active
    row_dt = [
        data.get("Summary").get("1yTargetEst"),
        data.get("Profile").get("fullTimeEmployees"),

        data.get("Statistics").get("averageDailyVolume3Month"),
        data.get("Statistics").get("shares_short"),
        data.get("Statistics").get("shortPercentOfFloat"),
        data.get("Statistics").get("sharesShortPriorMonth"),
        data.get("Statistics").get("trailingAnnualDividendRate"),

        data.get("Analysis").get("epsEstimate"),
        data.get("Analysis").get("epsActual"),
        data.get("Analysis").get("currentEstimate"),
        data.get("Analysis").get("7daysAgo"),
        data.get("Analysis").get("30daysAgo"),
        data.get("Analysis").get("currentYear"),
        data.get("Analysis").get("nextYear"),
        data.get("Analysis").get("next5Years"),
        data.get("Analysis").get("past5Years"),

        data.get("Profile").get("corporateGovernance"),

        data.get("Sustainability").get("totalEsg"),
        data.get("Sustainability").get("environmentScore"),
        data.get("Sustainability").get("socialScore"),
        data.get("Sustainability").get("governanceScore"),
        data.get("Sustainability").get("highestControversy"),

        data.get("Analysis").get("numberOfAnalysts").get("CurrentQtr"),
        data.get("Analysis").get("numberOfAnalysts").get("NextQtr"),
        data.get("Analysis").get("numberOfAnalysts").get("CurrentYear"),
        data.get("Analysis").get("numberOfAnalysts").get("NextYear"),

        data.get("Analysis").get("earnings_avg").get("CurrentQtr"),
        data.get("Analysis").get("earnings_avg").get("NextQtr"),
        data.get("Analysis").get("earnings_avg").get("CurrentYear"),
        data.get("Analysis").get("earnings_avg").get("NextYear"),

        data.get("Analysis").get("low").get("CurrentQtr"),
        data.get("Analysis").get("low").get("NextQtr"),
        data.get("Analysis").get("low").get("CurrentYear"),
        data.get("Analysis").get("low").get("NextYear"),

        data.get("Analysis").get("high").get("CurrentQtr"),
        data.get("Analysis").get("high").get("NextQtr"),
        data.get("Analysis").get("high").get("CurrentYear"),
        data.get("Analysis").get("high").get("NextYear"),

        data.get("Analysis").get("revenue_avg").get("CurrentQtr"),
        data.get("Analysis").get("revenue_avg").get("NextQtr"),
        data.get("Analysis").get("revenue_avg").get("CurrentYear"),
        data.get("Analysis").get("revenue_avg").get("NextYear"),

        data.get("Analysis").get("salesGrowth").get("CurrentQtr"),
        data.get("Analysis").get("salesGrowth").get("NextQtr"),
        data.get("Analysis").get("salesGrowth").get("CurrentYear"),
        data.get("Analysis").get("salesGrowth").get("NextYear")
    ]

    for i in range(len(row_dt)):
        sheet.cell(row=row, column=i + 67).value = row_dt[i]

    book.save("RESULTS/result.xlsx")
    book.close()


def get_saver():
    con = sqlite3.connect("all_analyst.db")
    cur = con.cursor()
    data = cur.execute("SELECT * FROM saver_yh").fetchall()

    data = sorted(data, key=lambda k: k[0])
    con.close()
    return data


def saver(sav_data):
    book = openpyxl.Workbook()
    sv = book.active

    saver_yh = ['Ticker', 'shares_short', 'epsEstimate', 'oper_date']

    for j in range(4):
        cell_header(sv.cell(row=1, column=j + 1), saver_yh[j], True)

    for i in range(len(sav_data)):
        for j in range(4):
            sv.cell(row=i + 2, column=j + 1).value = sav_data[i][j]

    book.save("RESULTS/Save_YH.xlsx")
    book.close()


def get_yahoo_cookie():
    cookie = None

    response = requests.get(
        "https://fc.yahoo.com", headers=headers, allow_redirects=True
    )

    if not response.cookies:
        return "NO_COOKIES"

    cookie = list(response.cookies)[0]

    return cookie


def get_yahoo_crumb(cookie):
    crumb = None

    crumb_response = requests.get(
        "https://query1.finance.yahoo.com/v1/test/getcrumb",
        headers=headers,
        cookies={cookie.name: cookie.value},
        allow_redirects=True,
    )
    crumb = crumb_response.text

    if crumb is None:
        return "NO_CRUMB"

    return crumb


def del_all():
    con = sqlite3.connect("all_analyst.db")
    cur = con.cursor()

    cur.execute("DELETE FROM yahoo")
    cur.execute("DELETE FROM yahoo_an")
    con.commit()
    con.close()


def save_data(ticker, data):
    summary = data.get("Summary")
    profile = data.get("Profile")
    statistics = data.get("Statistics")
    analysis = data.get("Analysis")
    sustainability = data.get("Sustainability")

    con = sqlite3.connect("all_analyst.db")
    cur = con.cursor()

    r = f"""INSERT INTO saver_yh VALUES
                    ({'?' + ',?' * 3});"""
    cur.execute(r, (ticker, statistics.get("shares_short"), analysis.get('epsEstimate'), date.today()))
    con.commit()

    r = f"""INSERT INTO yahoo VALUES
                        ({'?' + ',?' * 21});"""
    cur.execute(r, (ticker, summary.get('1yTargetEst'),
                    profile.get('fullTimeEmployees'),
                    statistics.get('averageDailyVolume3Month'),
                    statistics.get('shortPercentOfFloat'),
                    statistics.get('sharesShortPriorMonth'),
                    statistics.get('trailingAnnualDividendRate'),
                    analysis.get('epsEstimate'),
                    analysis.get('epsActual'),
                    analysis.get('currentEstimate'),
                    analysis.get('7daysAgo'),
                    analysis.get('30daysAgo'),
                    analysis.get('currentYear'),
                    analysis.get('nextYear'),
                    analysis.get('next5Years'),
                    analysis.get('past5Years'),
                    profile.get('corporateGovernance'),
                    sustainability.get('totalEsg'),
                    sustainability.get('environmentScore'),
                    sustainability.get('socialScore'),
                    sustainability.get('governanceScore'),
                    sustainability.get('highestControversy')))
    con.commit()

    r = f"""INSERT INTO yahoo_an VALUES
                    ({'?' + ',?' * 24});"""

    cur.execute(r, (ticker, analysis.get('numberOfAnalysts').get('CurrentQtr'),
                    analysis.get('numberOfAnalysts').get('NextQtr'),
                    analysis.get('numberOfAnalysts').get('NextQtr'),
                    analysis.get('numberOfAnalysts').get('NextQtr'),
                    analysis.get('earnings_avg').get('CurrentQtr'),
                    analysis.get('earnings_avg').get('NextQtr'),
                    analysis.get('earnings_avg').get('NextQtr'),
                    analysis.get('earnings_avg').get('NextQtr'),
                    analysis.get('low').get('CurrentQtr'),
                    analysis.get('low').get('NextQtr'),
                    analysis.get('low').get('NextQtr'),
                    analysis.get('low').get('NextQtr'),
                    analysis.get('high').get('CurrentQtr'),
                    analysis.get('high').get('NextQtr'),
                    analysis.get('high').get('NextQtr'),
                    analysis.get('high').get('NextQtr'),
                    analysis.get('revenue_avg').get('CurrentQtr'),
                    analysis.get('revenue_avg').get('NextQtr'),
                    analysis.get('revenue_avg').get('NextQtr'),
                    analysis.get('revenue_avg').get('NextQtr'),
                    analysis.get('salesGrowth').get('CurrentQtr'),
                    analysis.get('salesGrowth').get('NextQtr'),
                    analysis.get('salesGrowth').get('NextQtr'),
                    analysis.get('salesGrowth').get('NextQtr')))
    con.commit()
    con.close()


def main(appl_data):
    # if True:
    try:
        print()
        print("-=-=- YAHOO -=-=-")

        with open("config.txt") as f:
            data = f.readlines()
            YH_TIMEOUT = data[4].strip().split("=")[1].replace('"', '')
        # print([YH_TIMEOUT])

        row = 5
        del_all()
        for ticker in appl_data:
            if (row - 5) % COO == 0:
                cookie = get_yahoo_cookie()
                if cookie == "NO_COOKIES":
                    time.sleep(5)
                    cookie = get_yahoo_cookie()
                    if cookie == "NO_COOKIES":
                        print("COOKIE ERROR")
                        return
                crumb = get_yahoo_crumb(cookie)
                if crumb == "NO_CRUMB":
                    time.sleep(5)
                    crumb = get_yahoo_crumb(cookie)
                    if crumb == "NO_CRUMB":
                        print("CRUMB ERROR")
                        return
            inform = Parser(ticker, int(YH_TIMEOUT), cookie, crumb)

            # with open(f"yahoo/JSON/result_{ticker}.json", "w") as f:
            #     json.dump(inform.data, f)
            # with open(f"yahoo/JSON/result_{ticker}.json") as f:
            #     data = json.load(f)

            save_data(ticker, inform.data)
            to_xlsx(inform.data, row)
            row += 1
        sav_data = get_saver()
        saver(sav_data)
    except Exception:
        print("BIG ERROR YAHOO")

# ADMP
# ADMS
# ADS
# AEGN
# AENZ