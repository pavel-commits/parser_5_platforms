import time
from datetime import date
import requests
import json
import sqlite3

import openpyxl
from openpyxl.styles import PatternFill, Font

COLOR = "00FF00"


def get_data(USERNAME, PASSWORD, SCREEN_ID, SAVED_SCREEN_NAME):
    session = requests.Session()
    headers = {
        "Origin": "https://www.zacks.com",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.6 Safari/605.1.15",
        "Referer": "https://www.zacks.com/"
    }
    data = {
        "force_login": "true",
        "username": USERNAME,
        "password": PASSWORD
    }

    response_ses = session.post("https://www.zacks.com/", data=data, headers=headers)
    print("zacks_screener: 1", end=" ")

    response_scr = session.get(
        "https://screener-api.zacks.com/?scr_type=stock&c_id=zacks&c_key=0675466c5b74cfac34f6be7dc37d4fe6a008e212e2ef73bdcd7e9f1f9a9bd377&ecv=xQTM4AzMycDO&ref=screening",
        headers=headers)
    print("2", end=" ")

    headers[
        "Referer"] = "https://screener-api.zacks.com/?scr_type=stock&c_id=zacks&c_key=0675466c5b74cfac34f6be7dc37d4fe6a008e212e2ef73bdcd7e9f1f9a9bd377&ecv=xQTM4AzMycDO&ref=screening"

    response_my_screen = session.get("https://screener-api.zacks.com/myscreen.php?screen_type=1&_=1707567926132",
                                     headers=headers)
    print("3", end=" ")

    resp_load = session.post("https://screener-api.zacks.com/loadscreen.php", headers=headers, data={
        "screen_id": SCREEN_ID,
        "tab_id": "1"
    })
    print("4", end=" ")

    data1 = {
        "screen_id": SCREEN_ID,
        "mode": "runsavedscreen",
        "tab_id": "1"
    }
    response_get_dt1 = session.post("https://screener-api.zacks.com/getrunscreendata.php", headers=headers, data=data1)
    print("5", end=" ")

    resp_cr = session.get(
        "https://screener-api.zacks.com/getcriteriadata.php?category_id=10000&loadtype=A&category_name=Popular%20Criteria&_=1707568997563",
        headers=headers)
    print("6", end=" ")

    data2 = {
        "is_only_matches": "1",
        "is_premium_exists": "0",
        "is_edit_view": "0",
        "saved_screen_name": SAVED_SCREEN_NAME,
        "tab_id": "1",
        "start_page": "1",
        "no_of_rec": "15",
        "sort_col": "2",
        "sort_type": "ASC",
        "p_items[]": "12010",
        "p_item_name[]": "Market Cap (mil)",
        "p_item_key[]": "8",
        "operator[]": "6",
        "value[]": "10000",
        "config_params": "[]",
        "load_scr_id": SCREEN_ID
    }
    response_get_dt2 = session.post("https://screener-api.zacks.com/getrunscreendata.php", headers=headers, data=data2)
    print("7", end=" ")

    response_export = session.get("https://screener-api.zacks.com/export.php", headers=headers)
    print("8 ==> response")
    # with open("zacks_screener/response_export.txt", "w") as f:
    #     f.write(response_export.text)
    return response_export.text


def save_data(data):
    con = sqlite3.connect("all_analyst.db")
    cur = con.cursor()

    text = data.split("\n")

    text = text[1:-1]
    cur.execute("DELETE FROM zacks_screener")
    con.commit()

    r = f"""INSERT INTO zacks_screener VALUES
                ({'?' + ',?' * 18});"""

    for row in text:
        sp = row.strip().replace('""', '"0"')[1:-1].split('","')
        # print(sp)
        cur.execute(r, (sp[1], float(sp[2]), int(sp[3]), int(sp[4]), float(sp[5]), float(sp[6]), float(sp[7]),
                        float(sp[8]), float(sp[9]), float(sp[10]), float(sp[11]), float(sp[12]), int(sp[13]),
                        float(sp[14]), float(sp[15]), int(sp[16]), int(sp[17]), float(sp[18]), float(sp[19])))
        con.commit()
    con.close()


def cell_header(cell, text, color=False):
    cell.value = text
    cell.font = Font(bold=True)

    if color:
        cell.fill = PatternFill(start_color=COLOR, fill_type='solid')


def txt_to_dict(appl_data, data):
    dict_ret = {ticker: [None] * 18 for ticker in appl_data}
    text = data.split("\n")

    for row in text[1:-1]:
        sp = row.strip().replace('""', '"0"')[1:-1].split('","')
        if sp[1] in dict_ret:
            row_sp = [float(sp[2]), int(sp[3]), int(sp[4]), float(sp[5]), float(sp[6]), float(sp[7]),
                      float(sp[8]), float(sp[9]), float(sp[10]), float(sp[11]), float(sp[12]), int(sp[13]),
                      float(sp[14]), float(sp[15]), int(sp[16]), int(sp[17]), float(sp[18]), float(sp[19])]
            dict_ret[sp[1]] = row_sp
    return dict_ret


def to_xlsx(data):
    book = openpyxl.open("RESULTS/result.xlsx")
    sheet = book.active

    row_n = 5
    for ticker, val in data.items():

        for i in range(18):
            sheet.cell(row=row_n, column=i + 113).value = val[i]
        row_n += 1
    book.save("RESULTS/result.xlsx")
    book.close()


def main(appl_data):
    # if True:
    try:
        print()
        print("-=-=- ZACKS SCREENER -=-=-")

        with open("config.txt") as f:
            data = f.readlines()
            USERNAME = data[0].strip().split("=")[1].replace('"', '')
            PASSWORD = data[1].strip().split("=")[1].replace('"', '')
            SCREEN_ID = data[2].strip().split("=")[1].replace('"', '')
            SAVED_SCREEN_NAME = data[3].strip().split("=")[1].replace('"', '')
        # print([USERNAME, PASSWORD, SCREEN_ID, SAVED_SCREEN_NAME])

        data = get_data(USERNAME, PASSWORD, SCREEN_ID, SAVED_SCREEN_NAME)
        # with open("screener.json", "w") as f:
        #     json.dump(data, f)
        # data = open("zacks_screener/response_export.txt").read().strip()
        save_data(data)
        data = txt_to_dict(appl_data, data)
        to_xlsx(data)
    except Exception:
        print("ERROR, RESTART:")
        time.sleep(20)
        try:
            main(appl_data)
        except Exception:
            print("BIG ERROR ZACKS_SCREENER")
