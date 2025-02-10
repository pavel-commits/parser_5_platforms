import asyncio
import pandas as pd
from datetime import datetime, date
import requests
import time
import sqlite3

import openpyxl
from openpyxl.styles import PatternFill, Font

COLOR = "00FF00"


def cell_header(cell, text, color=False):
    cell.value = text
    cell.font = Font(bold=True)

    if color:
        cell.fill = PatternFill(start_color=COLOR, fill_type='solid')


weight_multiplier_dict = {
    "Bank of America": 1.5,
    "JPMorgan Chase & Co.": 1.5,
    "Evercore ISI": 1.5,
    "Morgan Stanley": 1.25,
    "Wolfe Research": 1.25,
    "Barclays": 1.25,
    "Jefferies Financial Group": 1.25,
    "Credit Suisse Group": 1.25,
    "Wells Fargo & Company": 1.25
}
today = datetime.now()
formatted_date = today.strftime("%Y%m%d")
output_file = f'Save_FMP.xlsx'


def calc_weight_multiplier(cell):
    return weight_multiplier_dict.get(cell, 1)


def get_price_targets(symbol):
    api_url = f'https://financialmodelingprep.com/api/v4/price-target?symbol={symbol}&apikey=e6a5965b93308e9e0e5e6bab76dbbda7'
    try:
        response = requests.get(api_url)
        response.raise_for_status()
        data = response.json()

        if data:
            return data
    except requests.exceptions.HTTPError as err:
        if response.status_code == 429:
            print(f"Rate limit exceeded. Waiting for 20 seconds.")
            time.sleep(20)
    return None


def save_data(symbol, averagepricetarget, rec2m):
    con = sqlite3.connect("all_analyst.db")
    cur = con.cursor()
    r = f"""INSERT INTO saver_fmp VALUES
                ({'?' + ',?' * 3});"""
    cur.execute(r, (symbol, averagepricetarget, rec2m, date.today()))
    con.commit()
    con.close()


def to_xlsx_sav(sav_data):
    book = openpyxl.Workbook()
    sheet = book.active

    saver_fmp = ['Ticker', 'AvgPriceTarget', 'Rec 2m', 'oper_date']
    for j in range(4):
        cell_header(sheet.cell(row=1, column=j + 1), saver_fmp[j], True)

    for i in range(len(sav_data)):
        for j in range(4):
            sheet.cell(row=i + 2, column=j + 1).value = sav_data[i][j]

    book.save(f"RESULTS/{output_file}")
    book.close()


def to_xlsx(row, averagepricetarget, rec2m):
    book = openpyxl.open("RESULTS/result.xlsx")
    sheet = book.active

    sheet.cell(row=row, column=2).value = averagepricetarget
    sheet.cell(row=row, column=3).value = rec2m

    book.save("RESULTS/result.xlsx")
    book.close()


def to_xlsx_tod(symbol, averagepricetarget, rec2m):
    book = openpyxl.open("RESULTS/result.xlsx")
    sheet = book.active

    i = 5
    while sheet[f"A{i}"].value:
        if sheet[f"A{i}"].value == symbol:
            break
        i += 1
    sheet.cell(row=i, column=2).value = averagepricetarget
    sheet.cell(row=i, column=3).value = rec2m

    book.save("RESULTS/result.xlsx")
    book.close()


def get_saver():
    con = sqlite3.connect("all_analyst.db")
    cur = con.cursor()
    data = cur.execute("SELECT * FROM saver_fmp").fetchall()

    data = sorted(data, key=lambda k: k[0])
    con.close()
    return data


async def compute_avg_price_target(today_appl):
    try:
        symbols_df = pd.DataFrame(today_appl, columns=['Symbol'])
        for i in symbols_df.index:
            symbol = symbols_df.at[i, 'Symbol']
            print(symbol)
            averagepricetarget, rec2m = None, None
            api_responses = get_price_targets(symbol)
            if api_responses:
                api_df = pd.DataFrame(api_responses,
                                      columns=['symbol', 'publishedDate', 'analystName', 'analystCompany', 'priceTarget'])
                analyst_df_filtered = api_df[['symbol', 'publishedDate', 'analystName', 'priceTarget']]

                totalweight = 0
                total = 0
                rec2m = 0

                for j in analyst_df_filtered.index:
                    d = analyst_df_filtered.at[j, 'publishedDate'][:10]
                    sd = datetime.strptime(d, '%Y-%m-%d')
                    pricetarget = analyst_df_filtered.at[j, 'priceTarget']
                    weight = max((sd - today).days + 180, 0)
                    weightmultiplier = calc_weight_multiplier(analyst_df_filtered.at[j, 'analystName'])
                    weightadjcell = weight * weightmultiplier
                    totalcell = weightadjcell * pricetarget

                    if (today - sd).days < 60:
                        rec2m += 1

                    totalweight += weightadjcell
                    total += totalcell

                if totalweight != 0:
                    averagepricetarget = total / totalweight
                else:
                    averagepricetarget = None

                symbols_df.at[i, 'AvgPriceTarget'] = averagepricetarget
                symbols_df.at[i, 'rec2m'] = rec2m
                save_data(symbol, averagepricetarget, rec2m)
            to_xlsx_tod(symbol, averagepricetarget, rec2m)
        # with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        #     symbols_df.to_excel(writer, sheet_name='Sheet1', index=False)

    except Exception as e:
        print(f"Error: {e}")


def main(today_appl):
    # if True:
    try:
        print()
        print("-=-=- FMP -=-=-")
        loop = asyncio.get_event_loop()
        loop.run_until_complete(compute_avg_price_target(today_appl))
        loop.close()

        sav_data = get_saver()
        to_xlsx_sav(sav_data)
    except Exception:
        print("BIG ERROR FMP")
