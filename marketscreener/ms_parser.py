import time
from datetime import date

import sqlite3
import requests
from bs4 import BeautifulSoup
import json

import openpyxl
from openpyxl.styles import PatternFill, Font

COLOR = "00FF00"
headers = {
    "Accept": "*/*",
    "Content-Type": "text/plain",
    "Accept-Language": "ru",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.6 Safari/605.1.15",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive"
}
DATABASE = json.load(open("marketscreener/DATABASE.json"))
NMB = ["2017 Q1", "2017 Q2", "2017 Q3", "2017 Q4", "2018 Q1", "2018 Q2", "2018 Q3", "2018 Q4", "2019 Q1", "2019 Q2",
       "2019 Q3", "2019 Q4", "2020 Q1", "2020 Q2", "2020 Q3", "2020 Q4", "2021 Q1", "2021 Q2", "2021 Q3", "2021 Q4",
       "2022 Q1", "2022 Q2", "2022 Q3", "2022 Q4", "2023 Q1", "2023 Q2", "2023 Q3", "2023 Q4", "2024 Q1", "2024 Q2",
       "2024 Q3", "2024 Q4", "2025 Q1", "2025 Q2", "2025 Q3", "2025 Q4"]


def make_url(ticker):
    try:
        global DATABASE

        url = DATABASE.get(ticker)
        if url:
            return url
        try:
            google_resp = requests.get(f"https://www.google.com/search?q=marketscreener+{ticker}", headers=headers)
        except Exception as e:
            print(e)
            print(ticker, "Google retry...")
            time.sleep(15)
            try:
                google_resp = requests.get(f"https://www.google.com/search?q=marketscreener+{ticker}", headers=headers)
            except:
                DATABASE[ticker] = None
                time.sleep(15)
                return url

        # with open(f"google{ticker}.html", "w") as f:
        #     f.write(google_resp.text)
        # with open(f"google{ticker}.html") as f:
        #     soup = BeautifulSoup(f, "html.parser")

        soup = BeautifulSoup(google_resp.text, "html.parser")
        url = soup.find("a", {"jsname": "UWckNb"})["href"]
        if url.startswith("https://www.marketscreener.com/quote/stock/"):
            while not url[-2].isdigit():
                url = "/".join(url.split("/")[:-2]) + "/"
            url += 'finances/'
            DATABASE[ticker] = url
        else:
            DATABASE[ticker] = None

        # json.dump(DATABASE, open("marketscreener/DATABASE.json", "w"))
        return url
    except Exception as e:
        return None


def parser(ticker, url, NOW):
    try:
        print(ticker)
        data = {
            "val": ["", ""],
            "Fiscal Period": None,
            "bsTable_ise": {},
            "iseTableQ": {}
        }
        if url is None:
            return data
        try:
            response = requests.get(url, headers=headers)
        except Exception as e:
            print(e)
            print("Retry...")
            time.sleep(10)
            try:
                response = requests.get(url, headers=headers)
            except:
                return data

        # with open(f"marketscreener/HTML/{ticker}.html", "w") as f:
        #     f.write(response.text)

        # with open(f"marketscreener/HTML/{ticker}.html") as f:
        #     soup = BeautifulSoup(f.read(), 'html.parser')

        soup = BeautifulSoup(response.text, 'html.parser')

        bsTable = soup.find("table", {"id": "bsTable"})
        iseTableA = soup.find("table", {"id": "iseTableA"})
        iseTableQ = soup.find("table", {"id": "iseTableQ"})

        if bsTable:
            try:
                trs = bsTable.find_all("tr")
                if trs:
                    try:
                        per = trs[0].find("i").text.split(":")[-1].strip()
                    except:
                        per = ""
                    data["Fiscal Period"] = per
                    ths = [th.text.replace("\n", "").strip() for th in trs[0].find_all("th")[1:]]
                    if "2024" in ths and "2025" in ths:
                        s4 = ths.index("2024")
                        trs = trs[1:]
                        i = 0
                        spis = ['Net Debt', 'Net Cash position', 'Leverage (Debt/EBITDA)', 'Free Cash Flow',
                                "ROE (net income / shareholders' equity)", 'ROA (Net income/ Total Assets)', 'Assets',
                                'Book Value Per Share', 'Cash Flow per Share', 'Capex', 'Capex / Sales',
                                'Announcement Date']
                        for tr in trs:
                            tds = tr.find_all("td")
                            name = tds[0].text.strip().replace("\n", "")
                            if name[-1].isdigit():
                                name = name[:-1].strip()
                            if tds:
                                if name != spis[i]:
                                    print("OOOOOOOOOOH NOOOOOOOAAAAAAAH")
                                tds = [td.text.strip().replace("\n", "").replace(",", "").replace("-", "0") for td in
                                       tds[1:]][s4:s4 + 2]
                                data["bsTable_ise"][spis[i]] = tds
                            i += 1
                        sp = soup.find("span", class_="pr-5").text.split()
                        data["val"] = [sp[0], sp[-1]]
                    elif "2024" in ths:
                        s4 = ths.index("2024")
                        trs = trs[1:]
                        i = 0
                        spis = ['Net Debt', 'Net Cash position', 'Leverage (Debt/EBITDA)', 'Free Cash Flow',
                                "ROE (net income / shareholders' equity)", 'ROA (Net income/ Total Assets)', 'Assets',
                                'Book Value Per Share', 'Cash Flow per Share', 'Capex', 'Capex / Sales',
                                'Announcement Date']
                        for tr in trs:
                            tds = tr.find_all("td")
                            name = tds[0].text.strip().replace("\n", "")
                            if name[-1].isdigit():
                                name = name[:-1].strip()

                            if tds:
                                if name != spis[i]:
                                    print("OOOOOOOOOOH NOOOOOOOAAAAAAAH")
                                tds = [td.text.strip().replace("\n", "").replace(",", "").replace("-", "0") for td in
                                       tds[1:]]
                                data["bsTable_ise"][spis[i]] = [tds[s4], "0"]
                            i += 1
                        sp = soup.find("span", class_="pr-5").text.split()
                        data["val"] = [sp[0], sp[-1]]
                    elif "2025" in ths:
                        s5 = ths.index("2025")
                        trs = trs[1:]
                        i = 0
                        spis = ['Net Debt', 'Net Cash position', 'Leverage (Debt/EBITDA)', 'Free Cash Flow',
                                "ROE (net income / shareholders' equity)", 'ROA (Net income/ Total Assets)', 'Assets',
                                'Book Value Per Share', 'Cash Flow per Share', 'Capex', 'Capex / Sales',
                                'Announcement Date']
                        for tr in trs:
                            tds = tr.find_all("td")
                            name = tds[0].text.strip().replace("\n", "")
                            if name[-1].isdigit():
                                name = name[:-1].strip()

                            if tds:
                                if name != spis[i]:
                                    print("OOOOOOOOOOH NOOOOOOOAAAAAAAH")
                                tds = [td.text.strip().replace("\n", "").replace(",", "").replace("-", "0") for td in
                                       tds[1:]]
                                data["bsTable_ise"][spis[i]] = ["0", tds[s5]]
                            i += 1
                        sp = soup.find("span", class_="pr-5").text.split()
                        data["val"] = [sp[0], sp[-1]]
            except:
                pass
        if iseTableA:
            try:
                trs = iseTableA.find_all("tr")
                if trs:
                    ths = [th.text.replace("\n", "").strip() for th in trs[0].find_all("th")[1:]]
                    if "2024" in ths and "2025" in ths:
                        s4 = ths.index("2024")
                        trs = trs[1:]
                        i = 0
                        spis = ['Net sales', 'EBITDA', 'EBIT', 'Operating Margin', 'Earnings before Tax (EBT)',
                                'Net income',
                                'Net margin', 'EPS', 'Free Cash Flow', 'FCF margin', 'FCF Conversion (EBITDA)',
                                'FCF Conversion (Net income)', 'Dividend per Share', 'Announcement Date']
                        for tr in trs:
                            tds = tr.find_all("td")
                            name = tds[0].text.strip().replace("\n", "")
                            if name[-1].isdigit():
                                name = name[:-1].strip()
                            if tds:
                                if name != spis[i]:
                                    print("OOOOOOOOOOH NOOOOOOOAAAAAAAH")
                                tds = [td.text.strip().replace("\n", "").replace(",", "").replace("-", "0") for td in
                                       tds[1:]][s4:s4 + 2]
                                data["bsTable_ise"][spis[i]] = tds
                            i += 1
                    elif "2024" in ths:
                        s4 = ths.index("2024")
                        trs = trs[1:]
                        i = 0
                        spis = ['Net sales', 'EBITDA', 'EBIT', 'Operating Margin', 'Earnings before Tax (EBT)',
                                'Net income',
                                'Net margin', 'EPS', 'Free Cash Flow', 'FCF margin', 'FCF Conversion (EBITDA)',
                                'FCF Conversion (Net income)', 'Dividend per Share', 'Announcement Date']
                        for tr in trs:
                            tds = tr.find_all("td")
                            name = tds[0].text.strip().replace("\n", "")
                            if name[-1].isdigit():
                                name = name[:-1].strip()
                            if tds:
                                if name != spis[i]:
                                    print("OOOOOOOOOOH NOOOOOOOAAAAAAAH")
                                tds = [td.text.strip().replace("\n", "").replace(",", "").replace("-", "0") for td in
                                       tds[1:]]
                                data["bsTable_ise"][spis[i]] = [tds[s4], "0"]
                    elif "2025" in ths:
                        s5 = ths.index("2025")
                        trs = trs[1:]
                        i = 0
                        spis = ['Net sales', 'EBITDA', 'EBIT', 'Operating Margin', 'Earnings before Tax (EBT)',
                                'Net income', 'Net margin', 'EPS', 'Free Cash Flow', 'FCF margin',
                                'FCF Conversion (EBITDA)', 'FCF Conversion (Net income)', 'Dividend per Share',
                                'Announcement Date']
                        for tr in trs:
                            tds = tr.find_all("td")
                            name = tds[0].text.strip().replace("\n", "")
                            if name[-1].isdigit():
                                name = name[:-1].strip()
                            if tds:
                                if name != spis[i]:
                                    print("OOOOOOOOOOH NOOOOOOOAAAAAAAH")
                                tds = [td.text.strip().replace("\n", "").replace(",", "").replace("-", "0") for td in
                                       tds[1:]]
                                data["bsTable_ise"][spis[i]] = ["0", tds[s5]]
            except:
                pass
        if iseTableQ:
            try:
                trs = iseTableQ.find_all("tr")
                if trs:
                    ths = [th.text.replace("\n", "").strip().replace("S", "Q") for th in trs[0].find_all("th")[1:]]
                    if NOW in ths:
                        s_now = ths.index(NOW)
                        trs = trs[1:]
                        i = 0
                        spis = ['Net sales', 'EBITDA', 'EBIT', 'Operating Margin', 'Earnings before Tax (EBT)',
                                'Net income', 'Net margin', 'EPS', 'Dividend per Share', 'Announcement Date']
                        for tr in trs:
                            tds = tr.find_all("td")
                            name = tds[0].text.strip().replace("\n", "")
                            if name[-1].isdigit():
                                name = name[:-1].strip()
                            if tds:
                                if name != spis[i]:
                                    print("OOOOOOOOOOH NOOOOOOOAAAAAAAH")
                                tds = [td.text.strip().replace("\n", "").replace(",", "").replace("-", "0") for td in
                                       tds[1:]]
                                data["iseTableQ"][spis[i]] = tds[s_now:s_now + 4]
                                len_sp = len(data["iseTableQ"][spis[i]])
                                if len_sp < 4:
                                    for _ in range(4 - len_sp):
                                        data["iseTableQ"][spis[i]].append("0.0")
                            i += 1
            except:
                pass
        # json.dump(data, open(f"marketscreener/JSON/{ticker}.json", "w"))
        return data
    except Exception:
        data = {
            "val": ["", ""],
            "Fiscal Period": None,
            "bsTable_ise": {},
            "iseTableQ": {}
        }
        return data


class Xlsx:
    def cell_header(self, cell, text, color=False):
        cell.value = text
        cell.font = Font(bold=True)
        # cell.alignment = Alignment(horizontal="center")

        if color:
            cell.fill = self.redFill

    def __init__(self):
        self.redFill = PatternFill(start_color=COLOR, fill_type='solid')
        # self.row = 5

    def saver(self, sav_data):
        book = openpyxl.Workbook()
        sv = book.active

        saver_ms = ['Ticker', '2024_Book Value Per Share', '2024_Net sales', '2024_EBITDA', '2024_EBIT', '2024_Net income',
                    '2024_EPS', '2025_Book Value Per Share', '2025_Net sales', '2025_EBITDA', '2025_EBIT',
                    '2025_Net income', '2025_EPS', '2024_Q1_EPS', '2024_Q2_EPS', 'oper_date']
        for j in range(16):
            self.cell_header(sv.cell(row=1, column=j + 1), saver_ms[j], True)

        for i in range(len(sav_data)):
            for j in range(16):
                sv.cell(row=i + 2, column=j + 1).value = sav_data[i][j]

        book.save("RESULTS/Save_MS.xlsx")
        book.close()

    def to_xlsx_tod(self, ticker, data):
        book = openpyxl.open("RESULTS/result.xlsx")
        sheet = book.active

        row = 5
        while sheet[f"A{row}"].value:
            if sheet[f"A{row}"].value == ticker:
                self.row = row
                break
            row += 1

        shift = 2
        # self.cell_header(sheet.cell(row=self.row, column=1), ticker)
        sheet.cell(row=self.row, column=shift + 2).value = data.get("Fiscal Period")
        sheet.cell(row=self.row, column=shift + 3).value = data.get("val", ["" * 2])[0]
        sheet.cell(row=self.row, column=shift + 4).value = data.get("val", ["" * 2])[1]

        if data.get("bsTable_ise"):
            for j in range(2):
                sheet.cell(row=self.row, column=shift + 5 + 14 * j).value = float(
                    data.get("bsTable_ise").get("Net Debt", ["0"] * 2)[j])
                sheet.cell(row=self.row, column=shift + 6 + 14 * j).value = float(
                    data.get("bsTable_ise").get("Net Cash position", ["0"] * 2)[j])
                sheet.cell(row=self.row, column=shift + 7 + 14 * j).value = float(
                    data.get("bsTable_ise").get("Assets", ["0"] * 2)[j])

                sheet.cell(row=self.row, column=shift + 8 + 14 * j).value = float(
                    data.get("bsTable_ise").get("Book Value Per Share", ["0"] * 2)[j])
                sheet.cell(row=self.row, column=shift + 9 + 14 * j).value = float(
                    data.get("bsTable_ise").get("Cash Flow per Share", ["0"] * 2)[j])
                sheet.cell(row=self.row, column=shift + 10 + 14 * j).value = float(
                    data.get("bsTable_ise").get("Capex", ["0"] * 2)[j])
                # -=-=-==-=-=
                sheet.cell(row=self.row, column=shift + 11 + 14 * j).value = float(
                    data.get("bsTable_ise").get("Net sales", ["0"] * 2)[j])
                sheet.cell(row=self.row, column=shift + 12 + 14 * j).value = float(
                    data.get("bsTable_ise").get("EBITDA", ["0"] * 2)[j])
                sheet.cell(row=self.row, column=shift + 13 + 14 * j).value = float(
                    data.get("bsTable_ise").get("EBIT", ["0"] * 2)[j])
                sheet.cell(row=self.row, column=shift + 14 + 14 * j).value = float(
                    data.get("bsTable_ise").get("Earnings before Tax (EBT)", ["0"] * 2)[j])
                sheet.cell(row=self.row, column=shift + 15 + 14 * j).value = float(
                    data.get("bsTable_ise").get("Net income", ["0"] * 2)[j])
                sheet.cell(row=self.row, column=shift + 16 + 14 * j).value = float(
                    data.get("bsTable_ise").get("EPS", ["0"] * 2)[j])
                sheet.cell(row=self.row, column=shift + 17 + 14 * j).value = float(
                    data.get("bsTable_ise").get("Free Cash Flow", ["0"] * 2)[j])
                sheet.cell(row=self.row, column=shift + 18 + 14 * j).value = float(
                    data.get("bsTable_ise").get("Dividend per Share", ["0"] * 2)[j])
        if data.get("iseTableQ"):
            for j in range(4):
                sheet.cell(row=self.row, column=shift + 33 + 8 * j).value = float(
                    data.get("iseTableQ").get("Net sales", ["0"] * 4)[j])
                sheet.cell(row=self.row, column=shift + 34 + 8 * j).value = float(
                    data.get("iseTableQ").get("EBITDA", ["0"] * 4)[j])
                sheet.cell(row=self.row, column=shift + 35 + 8 * j).value = float(
                    data.get("iseTableQ").get("EBIT", ["0"] * 4)[j])
                sheet.cell(row=self.row, column=shift + 36 + 8 * j).value = float(
                    data.get("iseTableQ").get("Earnings before Tax (EBT)", ["0"] * 4)[j])
                sheet.cell(row=self.row, column=shift + 37 + 8 * j).value = float(
                    data.get("iseTableQ").get("Net income", ["0"] * 4)[j])
                sheet.cell(row=self.row, column=shift + 38 + 8 * j).value = float(
                    data.get("iseTableQ").get("EPS", ["0"] * 4)[j])
                sheet.cell(row=self.row, column=shift + 39 + 8 * j).value = float(
                    data.get("iseTableQ").get("Dividend per Share", ["0"] * 4)[j])
                sheet.cell(row=self.row, column=shift + 40 + 8 * j).value = \
                    data.get("iseTableQ").get("Announcement Date", ["0"] * 4)[j]
        self.row += 1
        book.save("RESULTS/result.xlsx")
        book.close()


def get_saver():
    con = sqlite3.connect("all_analyst.db")
    cur = con.cursor()
    data = cur.execute("SELECT * FROM saver_ms").fetchall()

    data = sorted(data, key=lambda k: k[0])
    con.close()
    return data


def del_all(today_data):
    con = sqlite3.connect("all_analyst.db")
    cur = con.cursor()
    for ticker in today_data:
        cur.execute(f"DELETE FROM marketscreener WHERE Ticker='{ticker}'")
    con.commit()
    con.close()


def save_data(ticker, data):
    bs = data.get("bsTable_ise")
    ise = data.get("iseTableQ")

    con = sqlite3.connect("all_analyst.db")
    cur = con.cursor()

    r = f"""INSERT INTO saver_ms VALUES
                ({'?' + ',?' * 15});"""
    cur.execute(r, (ticker, bs.get("Book Value Per Share", [0, 0])[0],
                    bs.get("Net sales", [0, 0])[0],
                    bs.get("EBITDA", [0, 0])[0],
                    bs.get("EBIT", [0, 0])[0],
                    bs.get("Net income", [0, 0])[0],
                    bs.get("EPS", [0, 0])[0],
                    bs.get("Book Value Per Share", [0, 0])[1],
                    bs.get("Net sales", [0, 0])[1],
                    bs.get("EBITDA", [0, 0])[1],
                    bs.get("EBIT", [0, 0])[1],
                    bs.get("Net income", [0, 0])[1],
                    bs.get("EPS", [0, 0])[1],
                    ise.get("EPS", [0, 0])[0],
                    ise.get("EPS", [0, 0])[1],
                    date.today()))

    r = f"""INSERT INTO marketscreener VALUES
                ({'?' + ',?' * 49});"""

    cur.execute(r, (ticker,
                    data.get('Fiscal Period'),
                    data.get('val')[0],
                    data.get('val')[1],
                    bs.get('Net Debt', ["0", "0"])[0],
                    bs.get('Net Cash position', ["0", "0"])[0],
                    bs.get('Assets', ["0", "0"])[0],
                    bs.get('Cash Flow per Share', ["0", "0"])[0],
                    bs.get('Capex', ["0", "0"])[0],
                    bs.get('Earnings before Tax (EBT)', ["0", "0"])[0],
                    bs.get('Free Cash Flow', ["0", "0"])[0],
                    bs.get('Dividend per Share', ["0", "0"])[0],
                    bs.get('Net Debt', ["0", "0"])[1],
                    bs.get('Net Cash position', ["0", "0"])[1],
                    bs.get('Assets', ["0", "0"])[1],
                    bs.get('Cash Flow per Share', ["0", "0"])[1],
                    bs.get('Capex', ["0", "0"])[1],
                    bs.get('Earnings before Tax (EBT)', ["0", "0"])[1],
                    bs.get('Free Cash Flow', ["0", "0"])[1],
                    bs.get('Dividend per Share', ["0", "0"])[1],
                    ise.get('Net sales', ["0"] * 4)[0],
                    ise.get('EBITDA', ["0"] * 4)[0],
                    ise.get('EBIT', ["0"] * 4)[0],
                    ise.get('Earnings before Tax (EBT)', ["0"] * 4)[0],
                    ise.get('Net income', ["0"] * 4)[0],
                    ise.get('Dividend per Share', ["0"] * 4)[0],
                    ise.get('Announcement Date', ["0"] * 4)[0],
                    ise.get('Net sales', ["0"] * 4)[1],
                    ise.get('EBITDA', ["0"] * 4)[1],
                    ise.get('EBIT', ["0"] * 4)[1],
                    ise.get('Earnings before Tax (EBT)', ["0"] * 4)[1],
                    ise.get('Net income', ["0"] * 4)[1],
                    ise.get('Dividend per Share', ["0"] * 4)[1],
                    ise.get('Announcement Date', ["0"] * 4)[1],
                    ise.get('Net sales', ["0"] * 4)[2],
                    ise.get('EBITDA', ["0"] * 4)[2],
                    ise.get('EBIT', ["0"] * 4)[2],
                    ise.get('Earnings before Tax (EBT)', ["0"] * 4)[2],
                    ise.get('Net income', ["0"] * 4)[2],
                    ise.get('EPS', ["0"] * 4)[2],
                    ise.get('Dividend per Share', ["0"] * 4)[2],
                    ise.get('Announcement Date', ["0"] * 4)[2],
                    ise.get('Net sales', ["0"] * 4)[3],
                    ise.get('EBITDA', ["0"] * 4)[3],
                    ise.get('EBIT', ["0"] * 4)[3],
                    ise.get('Earnings before Tax (EBT)', ["0"] * 4)[3],
                    ise.get('Net income', ["0"] * 4)[3],
                    ise.get('EPS', ["0"] * 4)[3],
                    ise.get('Dividend per Share', ["0"] * 4)[3],
                    ise.get('Announcement Date', ["0"] * 4)[3]))
    con.commit()
    con.close()


def main(today_data):
    print()
    print("-=-=- MARKETSCREENER -=-=-")
    try:
        with open("config.txt") as f:
            data = f.readlines()
            MS_TIMEOUT = data[6].strip().split("=")[1].replace('"', '')
            NOW = data[7].strip().split("=")[1].replace('"', '')
        # print([MS_TIMEOUT, NOW])

        xlsx = Xlsx()
        del_all(today_data)
        for ticker in today_data:
            url = make_url(ticker)
            if url:
                data = parser(ticker, url, NOW)
                # with open(f"marketscreener/JSON/{ticker}.json") as f:
                #     data = json.load(f)
                # print(ticker)
                save_data(ticker, data)

                xlsx.to_xlsx_tod(ticker, data)
            time.sleep(int(MS_TIMEOUT))

        sav_data = get_saver()
        xlsx.saver(sav_data)
    except Exception:
        print("BIG ERROR MARKETSCREENER")
