from datetime import date, datetime
import openpyxl
import schedule

from yahoo import yahoo_parser
from marketscreener import ms_parser
from zacks import zacks_parser
from zacks_screener import screener_parser
from os import path, makedirs
from fmp import fmp_parser
from openpyxl.styles import PatternFill, Font

COLOR = "00FF00"
NMB = ["2017 Q1", "2017 Q2", "2017 Q3", "2017 Q4", "2018 Q1", "2018 Q2", "2018 Q3", "2018 Q4", "2019 Q1", "2019 Q2",
       "2019 Q3", "2019 Q4", "2020 Q1", "2020 Q2", "2020 Q3", "2020 Q4", "2021 Q1", "2021 Q2", "2021 Q3", "2021 Q4",
       "2022 Q1", "2022 Q2", "2022 Q3", "2022 Q4", "2023 Q1", "2023 Q2", "2023 Q3", "2023 Q4", "2024 Q1", "2024 Q2",
       "2024 Q3", "2024 Q4", "2025 Q1", "2025 Q2", "2025 Q3", "2025 Q4"]


def cell_header(cell, text, color=False):
    cell.value = text
    cell.font = Font(bold=True)

    if color:
        cell.fill = PatternFill(start_color=COLOR, fill_type='solid')


def appl_checker(appl_data):
    if not path.exists("RESULTS"):
        return True
    if not path.exists("RESULTS/result.xlsx"):
        return True

    book = openpyxl.open("RESULTS/result.xlsx")
    sheet = book.active

    i = 5
    appl_data_res = []

    while sheet[f"A{i}"].value:
        ticker, freq = sheet[f"A{i}"].value, sheet[f"B{i}"].value
        appl_data_res.append(ticker.upper())
        i += 1
    book.close()
    if not appl_data == appl_data_res:
        return True
    return False


def get_appl_data_xlsx():
    appl_data = []

    book = openpyxl.open("symbols.xlsx")
    sheet = book.active

    i = 1
    while sheet[f"A{i}"].value or sheet[f"B{i}"].value:
        ticker, freq = sheet[f"A{i}"].value, sheet[f"B{i}"].value
        # if not ticker or not freq or not freq in ["everywd", "tuefri", "tue"]:
        #     print(f"Error: {i} row of symbols.xlsx")
        #     return "ERROR"

        # appl_data.append((ticker.upper(), freq))
        if not ticker.upper() in appl_data:
            appl_data.append(ticker.upper())
        i += 1
    book.close()
    return appl_data


def get_today_appl(fmp_ms_parser_per):
    today_appl = []
    book = openpyxl.open("symbols.xlsx")
    sheet = book.active
    i = 1
    while sheet[f"A{i}"].value or sheet[f"B{i}"].value:
        ticker, freq = sheet[f"A{i}"].value, sheet[f"B{i}"].value

        if not ticker.upper() in today_appl and freq in fmp_ms_parser_per:
            today_appl.append(ticker.upper())
        i += 1
    book.close()
    return today_appl


def xlsx_main(appl_data, NOW):
    folder = "RESULTS"
    if not path.exists(folder):
        makedirs(folder)

    book = openpyxl.Workbook()
    sheet = book.active

    cell_header(sheet.cell(row=3, column=1), "Ticker")

    # -=-=-=-=-=-=-=-==-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=

    cell_header(sheet.cell(row=1, column=2), "FMP", True)
    # cell_header(sheet.cell(row=2, column=2), date.today())

    cell_header(sheet.cell(row=3, column=2), "AvgPriceTarget")
    cell_header(sheet.cell(row=3, column=3), "Rec 2m")

    # -=-=-=-=-=-=-=-==-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=

    shift = 3
    cell_header(sheet.cell(row=1, column=shift + 1), "MARKETSCREENER", True)

    # cell_header(sheet.cell(row=2, column=shift + 1), date.today())

    # cell_header(sheet.cell(row=3, column=1), "Tickers")
    cell_header(sheet.cell(row=3, column=shift + 1), "Fiscal period")
    cell_header(sheet.cell(row=3, column=shift + 2), "Currency")
    cell_header(sheet.cell(row=3, column=shift + 3), "Size")

    bsTable_ise = ['Net Debt', 'Net Cash position', 'Assets', 'Book Value Per Share', 'Cash Flow per Share',
                   'Capex', 'Net sales', 'EBITDA', 'EBIT', 'Earnings before Tax (EBT)', 'Net income', 'EPS',
                   'Free Cash Flow', 'Dividend per Share']
    iseTableQ = ['Net sales', 'EBITDA', 'EBIT', 'Earnings before Tax (EBT)', 'Net income', 'EPS',
                 'Dividend per Share', 'Announcement Date']
    i = 7
    for j in range(2):
        cell_header(sheet.cell(row=3, column=i), 2024 + j, color=True)
        for val in bsTable_ise:
            cell_header(sheet.cell(row=4, column=i), val)
            i += 1
    for j in range(NMB.index(NOW), NMB.index(NOW) + 4):
        cell_header(sheet.cell(row=3, column=i), NMB[j], color=True)
        for val in iseTableQ:
            cell_header(sheet.cell(row=4, column=i), val)
            i += 1

    # -=-=-=-=-=-=-=-==-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=

    shift = 66
    cell_header(sheet.cell(row=1, column=shift + 1), "YAHOO", True)
    # cell_header(sheet.cell(row=2, column=shift + 1), date.today())

    yahoo = ['1yTargetEst', 'fullTimeEmployees', 'averageDailyVolume3Month', 'shares_short',
             'shortPercentOfFloat', 'sharesShortPriorMonth', 'trailingAnnualDividendRate', 'epsEstimate',
             'epsActual', 'currentEstimate', '7daysAgo', '30daysAgo', 'currentYear', 'nextYear', 'next5Years',
             'past5Years', 'corporateGovernance', 'totalEsg', 'environmentScore', 'socialScore', 'governanceScore',
             'highestControversy']
    yahoo_an = ['numberOfAnalysts_Qtr_cur', 'numberOfAnalysts_Qtr_next', 'numberOfAnalysts_Year_cur',
                'numberOfAnalysts_Year_next', 'earnings_avg_Qtr_cur', 'earnings_avg_Qtr_next', 'earnings_avg_Year_cur',
                'earnings_avg_Year_next', 'low_Qtr_cur', 'low_Qtr_next', 'low_Year_cur', 'low_Year_next',
                'high_Qtr_cur', 'high_Qtr_next', 'high_Year_cur', 'high_Year_next', 'revenue_avg_Qtr_cur',
                'revenue_avg_Qtr_next', 'revenue_avg_Year_cur', 'revenue_avg_Year_next', 'salesGrowth_Qtr_cur',
                'salesGrowth_Qtr_next', 'salesGrowth_Year_cur', 'salesGrowth_Year_next']

    for i in range(len(yahoo)):
        cell_header(sheet.cell(row=3, column=shift + i + 1), yahoo[i])

    for i in range(len(yahoo_an)):
        cell_header(sheet.cell(row=3, column=shift + i + len(yahoo) + 1), yahoo_an[i])

    # -=-=-=-=-=-=-=-==-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=

    shift = 112
    cell_header(sheet.cell(row=1, column=shift + 1), "ZACKS SCREENER", True)
    # cell_header(sheet.cell(row=2, column=shift + 1), date.today())

    text = ['Market Cap (mil)', 'Last EPS Report Date (yyyymmdd)', 'Next EPS Report Date  (yyyymmdd)',
            '% Change Q1 Est. (4 weeks)', '% Change Q2 Est. (4 weeks)', '% Change F1 Est. (4 weeks)',
            '% Change F2 Est. (4 weeks)', 'Q1 Consensus Est. ', 'Q2 Consensus Est. (next fiscal Qtr)',
            'F1 Consensus Est.', 'F2 Consensus Est.', '# of Analysts in Q1 Consensus',
            'F(1) Consensus Sales Est. ($mil)', 'Q(1) Consensus Sales Est. ($mil)', '# of Analysts in F1 Consensus',
            '# of Analysts in F2 Consensus', 'St. Dev. Q1 / Q1 Consensus', 'St. Dev. F1 / F1 Consensus']
    i = 1
    for t in text:
        cell_header(sheet.cell(row=3, column=shift + i), t)
        i += 1

    # -=-=-=-=-=-=-=-==-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=

    shift = 130
    cell_header(sheet.cell(row=1, column=shift + 1), "ZACKS DETAILED", color=True)
    # cell_header(sheet.cell(row=2, column=shift + 1), date.today())

    # cell_header(sheet.cell(row=3, column=1), "Symbol")
    cell_header(sheet.cell(row=3, column=shift + 1), "Sales", color=True)
    cell_header(sheet.cell(row=3, column=shift + 5), "Earnings", color=True)

    cell_header(sheet.cell(row=4, column=shift + 1), "Current_high")
    cell_header(sheet.cell(row=4, column=shift + 2), "Current_low")
    cell_header(sheet.cell(row=4, column=shift + 3), "Next_high")
    cell_header(sheet.cell(row=4, column=shift + 4), "Next_low")

    cell_header(sheet.cell(row=4, column=shift + 5), "Current_high")
    cell_header(sheet.cell(row=4, column=shift + 6), "Current_low")
    cell_header(sheet.cell(row=4, column=shift + 7), "Next_high")
    cell_header(sheet.cell(row=4, column=shift + 8), "Next_low")

    book.save("RESULTS/result.xlsx")
    # -=-=-=-=-=-=-=-==-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=

    row = 5
    for ticker in appl_data:
        cell_header(sheet.cell(row=row, column=1), ticker)
        row += 1
    book.save("RESULTS/result.xlsx")
    book.close()


def clear_res(beg, end, len_ap, is_ff=False):
    book = openpyxl.open("RESULTS/result.xlsx")
    sheet = book.active

    for i in range(beg, end):
        for j in range(5, len_ap + 5):
            sheet.cell(row=j, column=i).value = ""

    cell_header(sheet.cell(row=2, column=beg), date.today())

    if is_ff:
        cell_header(sheet.cell(row=2, column=4), date.today())
    book.save("RESULTS/result.xlsx")
    book.close()


def clear_res_ms_fmp(today_appl, len_ap):
    book = openpyxl.open("RESULTS/result.xlsx")
    sheet = book.active

    for i in range(2, 67):
        for j in range(5, len_ap + 5):
            if sheet.cell(row=j, column=1).value in today_appl:
                sheet.cell(row=j, column=i).value = ""

    cell_header(sheet.cell(row=2, column=2), datetime.today())
    cell_header(sheet.cell(row=2, column=4), datetime.today())

    book.save("RESULTS/result.xlsx")
    book.close()


def act(today_appl, yahoo_parser_per, screener_parser_per, zacks_parser_per, appl_data):
    len_ap = len(appl_data)
    # print("today_appl:", today_appl)

    if today_appl == "all":
        clear_res(2, 4, len_ap, is_ff=True)
        # cell_header(sheet.cell(row=2, column=2), datetime.today())

        fmp_parser.main(appl_data)
        with open("log.txt", "a") as log_txt:
            log_txt.write(f"FMP {date.today()}\n")

        # cell_header(sheet.cell(row=2, column=4), datetime.today())

        ms_parser.main(appl_data)
        with open("log.txt", "a") as log_txt:
            log_txt.write(f"MARKETSCREENER {date.today()}\n")
    else:
        clear_res_ms_fmp(today_appl, len_ap)
        # cell_header(sheet.cell(row=2, column=2), datetime.today())

        fmp_parser.main(today_appl)
        with open("log.txt", "a") as log_txt:
            log_txt.write(f"FMP {date.today()}\n")

        # cell_header(sheet.cell(row=2, column=4), datetime.today())

        ms_parser.main(today_appl)
        with open("log.txt", "a") as log_txt:
            log_txt.write(f"MARKETSCREENER {date.today()}\n")

    if yahoo_parser_per:
        # cell_header(sheet.cell(row=2, column=67), datetime.today())

        clear_res(67, 113, len_ap)
        yahoo_parser.main(appl_data)
        with open("log.txt", "a") as log_txt:
            log_txt.write(f"YAHOO {date.today()}\n")

    if screener_parser_per:
        # cell_header(sheet.cell(row=2, column=113), datetime.today())

        clear_res(113, 131, len_ap)
        screener_parser.main(appl_data)
        with open("log.txt", "a") as log_txt:
            log_txt.write(f"ZACKS SCREENER {date.today()}\n")

    if zacks_parser_per:
        # cell_header(sheet.cell(row=2, column=131), datetime.today())

        clear_res(131, 139, len_ap)
        zacks_parser.main(appl_data)
        with open("log.txt", "a") as log_txt:
            log_txt.write(f"ZACKS.COM {date.today()}\n")


def do_it(screener_parser_per, fmp_ms_parser_per, yahoo_parser_per, zacks_parser_per):
    print(datetime.today())
    with open("config.txt") as f:
        data = f.readlines()
        NOW = data[7].strip().split("=")[1].replace('"', '')

    appl_data = get_appl_data_xlsx()
    today_appl = get_today_appl(fmp_ms_parser_per)

    if appl_checker(appl_data):
        print("UPDATE FILE")
        xlsx_main(appl_data, NOW)

        act(today_appl="all",
            yahoo_parser_per=True,
            screener_parser_per=True,
            zacks_parser_per=True,
            appl_data=appl_data)
    else:
        act(today_appl=today_appl,
            yahoo_parser_per=yahoo_parser_per,
            screener_parser_per=screener_parser_per,
            zacks_parser_per=zacks_parser_per,
            appl_data=appl_data)
    print()
    print("-=-=- DONE -=-=-")


def main():
    print("Script is active")
    schedule.every().monday.at("21:30").do(do_it, fmp_ms_parser_per=["everywd"],
                                           yahoo_parser_per=False,
                                           screener_parser_per=True,
                                           zacks_parser_per=False)

    schedule.every().tuesday.at("21:30").do(do_it, fmp_ms_parser_per=["everywd", "tue", "tuefri"],
                                            yahoo_parser_per=True,
                                            screener_parser_per=True,
                                            zacks_parser_per=True)

    schedule.every().wednesday.at("21:30").do(do_it, fmp_ms_parser_per=["everywd"],
                                              yahoo_parser_per=False,
                                              screener_parser_per=True,
                                              zacks_parser_per=False)

    schedule.every().thursday.at("21:30").do(do_it, fmp_ms_parser_per=["everywd", "tue", "tuefri"],
                                             yahoo_parser_per=False,
                                             screener_parser_per=True,
                                             zacks_parser_per=False)

    schedule.every().friday.at("21:30").do(do_it, fmp_ms_parser_per=["everywd", "tuefri"],
                                           yahoo_parser_per=False,
                                           screener_parser_per=True,
                                           zacks_parser_per=False)
    # -=-=-=-=-=-=-=-=-=-=-=-=-=-=

    while True:
        schedule.run_pending()


if __name__ == '__main__':
    main()
